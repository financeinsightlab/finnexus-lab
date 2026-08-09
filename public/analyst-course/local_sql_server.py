#!/usr/bin/env python3
"""SignalPath local-only server with a read-only SQLite execution endpoint.
No cloud service, credentials, or outbound network access is required.
Run: python3 local_sql_server.py
"""
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
import ast, base64, contextlib, csv, io, json, os, re, sqlite3, traceback
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
TABLE_FILES = {
    "accounts": ROOT / "datasets/sql-saas/accounts.csv",
    "subscriptions": ROOT / "datasets/sql-saas/subscriptions.csv",
    "product_events": ROOT / "datasets/sql-saas/product_events.csv",
    "support_tickets": ROOT / "datasets/sql-saas/support_tickets.csv",
}


def build_db():
    db = sqlite3.connect(":memory:")
    db.row_factory = sqlite3.Row
    for table, path in TABLE_FILES.items():
        with path.open(newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        columns = list(rows[0].keys())
        db.execute(f'CREATE TABLE "{table}" ({", ".join(f"\"{c}\" TEXT" for c in columns)})')
        db.executemany(
            f'INSERT INTO "{table}" ({", ".join(f"\"{c}\"" for c in columns)}) VALUES ({", ".join("?" for _ in columns)})',
            [[r[c] for c in columns] for r in rows],
        )
    return db


def valid_read_only(sql: str) -> tuple[bool, str]:
    cleaned = re.sub(r"--.*?$|/\*.*?\*/", "", sql, flags=re.MULTILINE | re.DOTALL).strip()
    if not cleaned:
        return False, "Enter a SQL query first."
    if ";" in cleaned.rstrip(";"):
        return False, "Only one read-only SQL statement is allowed."
    if not re.match(r"^(select|with|explain)\b", cleaned, re.I):
        return False, "This local lab accepts only SELECT, WITH, or EXPLAIN statements."
    blocked = r"\b(insert|update|delete|drop|alter|create|attach|detach|vacuum|pragma|replace)\b"
    if re.search(blocked, cleaned, re.I):
        return False, "Write operations and database commands are blocked in the local lab."
    return True, cleaned


def python_tables():
    customers = pd.read_csv(ROOT / "datasets/python-customer/customers.csv")
    transactions = pd.read_csv(ROOT / "datasets/python-customer/transactions.csv")
    return customers, transactions


def validate_python(code: str) -> tuple[bool, str]:
    if not code.strip():
        return False, "Enter Python code first."
    if len(code) > 12000:
        return False, "Keep a single learning cell below 12,000 characters."
    try:
        tree = ast.parse(code)
    except SyntaxError as exc:
        return False, f"Python syntax: {exc.msg} (line {exc.lineno})"
    banned_names = {"open", "exec", "eval", "compile", "globals", "locals", "vars", "input", "help", "__import__", "getattr", "setattr", "delattr"}
    allowed_imports = {"pandas", "numpy", "matplotlib.pyplot", "seaborn"}
    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            modules = [a.name for a in node.names] if isinstance(node, ast.Import) else [node.module or ""]
            if any(m not in allowed_imports for m in modules):
                return False, "Only pandas, numpy, matplotlib.pyplot and seaborn imports are allowed in this local learning cell."
        if isinstance(node, ast.Name) and (node.id in banned_names or "__" in node.id):
            return False, "This local learning cell blocks file, system, reflection and dynamic-execution functions."
        if isinstance(node, ast.Attribute) and "__" in node.attr:
            return False, "Double-underscore attributes are blocked in this local learning cell."
    return True, ""


def dataframe_payload(value):
    if isinstance(value, pd.DataFrame):
        sample = value.head(50).where(pd.notna(value.head(50)), None)
        return {"kind": "table", "columns": [str(c) for c in sample.columns], "rows": sample.astype(object).values.tolist(), "shape": list(value.shape)}
    if isinstance(value, pd.Series):
        sample = value.head(50).where(pd.notna(value.head(50)), None)
        return {"kind": "table", "columns": [str(value.name or "value")], "rows": [[v] for v in sample.astype(object).tolist()], "shape": [len(value), 1]}
    return {"kind": "text", "text": repr(value)[:4000]}


def execute_python(code: str):
    ok, error = validate_python(code)
    if not ok:
        return {"ok": False, "error": error}
    customers, transactions = python_tables()
    output = io.StringIO()
    safe_builtins = {"print": print, "len": len, "range": range, "min": min, "max": max, "sum": sum, "round": round, "abs": abs, "sorted": sorted, "list": list, "dict": dict, "set": set, "tuple": tuple, "str": str, "int": int, "float": float, "bool": bool, "enumerate": enumerate, "zip": zip}
    scope = {"__builtins__": safe_builtins, "pd": pd, "np": np, "customers": customers, "transactions": transactions}
    try:
        with contextlib.redirect_stdout(output):
            exec(compile(code, "<local-python-lab>", "exec"), scope, scope)
        value = scope.get("result")
        payload = dataframe_payload(value) if value is not None else {"kind": "text", "text": output.getvalue() or "Code ran successfully. Assign a DataFrame or Series to `result` to preview it here."}
        payload.update({"ok": True, "stdout": output.getvalue()})
        return payload
    except Exception as exc:
        return {"ok": False, "error": "Python: " + str(exc), "trace": traceback.format_exc(limit=2)}


def validate_workbook(encoded: str):
    try:
        raw = base64.b64decode(encoded, validate=True)
    except Exception:
        return {"ok": False, "error": "The uploaded file was not valid base64 workbook data."}
    if len(raw) > 12 * 1024 * 1024:
        return {"ok": False, "error": "Keep workbook below 12 MB for local validation."}
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=False, read_only=False)
    except Exception as exc:
        return {"ok": False, "error": f"Workbook could not be read: {exc}"}
    names = {name.lower().strip(): name for name in wb.sheetnames}
    required = ["raw_orders", "clean_orders", "analysis", "dashboard", "cleaning_log"]
    present = {key: key in names for key in required}
    formulas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append((ws.title, cell.coordinate, cell.value.upper()))
    formula_text = "\n".join(f[2] for f in formulas)
    checks = {
        "required_sheets": all(present.values()),
        "raw_data_present": wb[names["raw_orders"]].max_row > 1 if "raw_orders" in names else False,
        "formula_evidence": len(formulas) > 0,
        "sumifs_or_pivot_evidence": "SUMIFS(" in formula_text or "SUMIF(" in formula_text,
        "gross_profit_evidence": any("REVENUE" in x and "COST" in x and "-" in x for _, _, x in formulas),
        "margin_evidence": any("/" in x and ("REVENUE" in x or "SUMIFS" in x) for _, _, x in formulas),
    }
    score = round(sum(checks.values()) / len(checks) * 100)
    feedback = []
    if not checks["required_sheets"]:
        missing = [s for s, exists in present.items() if not exists]
        feedback.append("Add required worksheet(s): " + ", ".join(missing) + ".")
    if not checks["formula_evidence"]:
        feedback.append("Add auditable formulas in analysis or dashboard sheets.")
    if not checks["sumifs_or_pivot_evidence"]:
        feedback.append("Add a SUMIFS/SUMIF KPI formula or document pivot-table evidence in your workbook.")
    if not checks["gross_profit_evidence"]:
        feedback.append("Add a gross profit calculation using revenue minus cost.")
    if not checks["margin_evidence"]:
        feedback.append("Add a gross-margin calculation using aggregated profit divided by aggregated revenue.")
    if not feedback:
        feedback.append("Workbook structure and formula signals satisfy the local evidence check. Review business definitions and dashboard quality before publishing.")
    return {"ok": True, "score": score, "sheets": wb.sheetnames, "required": present, "formula_count": len(formulas), "formula_samples": [{"sheet": s, "cell": c, "formula": f} for s, c, f in formulas[:12]], "checks": checks, "feedback": feedback}


class Handler(SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_POST(self):
        if self.path not in {"/api/sql", "/api/python", "/api/spreadsheet"}:
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = json.loads(self.rfile.read(length).decode("utf-8"))
            if self.path == "/api/python":
                payload = execute_python(str(body.get("code", "")))
                self.respond(payload, 200 if payload.get("ok") else 400)
                return
            if self.path == "/api/spreadsheet":
                payload = validate_workbook(str(body.get("workbook", "")))
                self.respond(payload, 200 if payload.get("ok") else 400)
                return
            ok, sql = valid_read_only(str(body.get("sql", "")))
            if not ok:
                self.respond({"ok": False, "error": sql}, 400)
                return
            db = build_db()
            cursor = db.execute(sql)
            rows = cursor.fetchmany(200)
            columns = [d[0] for d in cursor.description] if cursor.description else []
            self.respond({"ok": True, "columns": columns, "rows": [[r[c] for c in columns] for r in rows], "row_limit": 200})
        except sqlite3.Error as exc:
            self.respond({"ok": False, "error": f"SQLite: {exc}"}, 400)
        except Exception as exc:
            self.respond({"ok": False, "error": f"Local lab error: {exc}"}, 500)

    def respond(self, payload, status=200):
        raw = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)


if __name__ == "__main__":
    os.chdir(ROOT)
    server = ThreadingHTTPServer(("0.0.0.0", 8001), Handler)
    print("SignalPath local SQL server running at http://0.0.0.0:8001")
    server.serve_forever()
